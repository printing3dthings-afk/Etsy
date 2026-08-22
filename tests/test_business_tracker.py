"""
Tests for the Business Tracker workbook (2026-07-18) — one .xlsx, multiple
tabs, generated fresh from live data on GET /api/business-tracker.xlsx.

Covers three layers:
  1. tools/business_tracker.py's build_workbook() as a pure module — fixture
     data in, verify sheet names/headers/values (hand-computed, same style
     as tests/test_cogs_status.py's fee-math verification).
  2. The orders-fetch dedup: _get_recent_orders_raw() must be the ONE place
     that calls EtsyAPIClient().get_orders() -- _search_orders() and
     _sales_by_listing_sync() must both reuse its cache, not each fetch
     independently (this was the actual duplication being closed).
  3. The endpoint itself via FastAPI TestClient — auth required, correct
     content-type, and the returned bytes are a real openable workbook.

Run: python tests/test_business_tracker.py
"""
import io
import json
import os
import sys
import tempfile
import traceback
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent

_tmp_db = tempfile.NamedTemporaryFile(prefix="frank_tracker_test_", suffix=".db", delete=False)
_tmp_db.close()
os.environ["DB_PATH"] = _tmp_db.name
os.environ.setdefault("APP_SECRET_TOKEN", "tracker-test-not-a-real-secret")

for p in (ROOT / "tools" / "api_server", ROOT / "tools"):
    sp = str(p)
    if sp not in sys.path:
        sys.path.insert(0, sp)

import main as server  # noqa: E402
import business_tracker  # noqa: E402
from openpyxl import load_workbook  # noqa: E402
from fastapi.testclient import TestClient  # noqa: E402

_failures: list[str] = []


def check(cond: bool, msg: str) -> None:
    if not cond:
        _failures.append(msg)


# ── fixture data, shared across module + endpoint tests ────────────────────
_LISTINGS = [{"listing_id": 111, "title": "Digital Planner 2026 Undated GoodNotes", "price": 14.99}]
_SALES = {111: 3}
_ORDERS_RAW = [{
    "receipt_id": 999, "name": "Jane Buyer", "create_timestamp": 1752000000,
    "grandtotal": {"amount": 1499, "divisor": 100},
    "transactions": [{"title": "Digital Planner 2026 Undated GoodNotes", "listing_id": 111, "quantity": 1}],
}]
_CATALOG = [{
    "product_id": "DP1026", "name": "Ultimate Digital Life Planner", "category": "digital_planner",
    "status": "active", "price": 14.99, "etsy_listing_id": "111", "last_updated": "2026-07-18",
    "note": "test fixture",
}]


# ── 1. business_tracker.build_workbook() as a pure module ──────────────────
def test_build_workbook_has_all_expected_sheets():
    buf = business_tracker.build_workbook(_LISTINGS, _SALES, _ORDERS_RAW, _CATALOG, server._estimate_listing_economics)
    check(isinstance(buf, io.BytesIO), f"build_workbook must return a BytesIO, got {type(buf)}")
    wb = load_workbook(buf)
    expected = {
        "Dashboard", "Products", "COGS & Profit", "Orders",
        "Physical Inventory", "Consumables & Reorder", "Suppliers",
        "Equipment & Assets", "Expense & Tax Tracker",
    }
    check(set(wb.sheetnames) == expected, f"expected sheets {expected}, got {set(wb.sheetnames)}")
    check(wb.sheetnames[0] == "Dashboard", f"Dashboard must be the first tab, got order: {wb.sheetnames}")


def test_products_sheet_reflects_the_live_catalog_not_a_hardcoded_snapshot():
    buf = business_tracker.build_workbook(_LISTINGS, _SALES, _ORDERS_RAW, _CATALOG, server._estimate_listing_economics)
    wb = load_workbook(buf)
    ws = wb["Products"]
    row = [c.value for c in ws[5]]  # header at row 4, first data row at row 5
    check(row[0] == "DP1026", f"Products sheet must reflect the passed-in catalog, got row: {row}")
    check(row[4] == 14.99, f"expected price 14.99 in Products row, got: {row}")


def test_cogs_profit_sheet_uses_real_fee_math_and_real_units_sold():
    buf = business_tracker.build_workbook(_LISTINGS, _SALES, _ORDERS_RAW, _CATALOG, server._estimate_listing_economics)
    wb = load_workbook(buf)
    ws = wb["COGS & Profit"]
    row = [c.value for c in ws[5]]
    expected_econ = server._estimate_listing_economics(14.99, "Digital Planner 2026 Undated GoodNotes")
    check(row[0] == 111, f"expected listing_id 111, got: {row}")
    check(row[4] == 3, f"expected recent_units_sold=3 (from the sales map), got: {row}")
    check(row[6] == expected_econ["cogs_estimate"], f"COGS estimate must match _estimate_listing_economics, got: {row}")
    check(row[8] == expected_econ["margin_pct"], f"margin %% must match _estimate_listing_economics, got: {row}")


def test_orders_sheet_is_live_receipt_data():
    buf = business_tracker.build_workbook(_LISTINGS, _SALES, _ORDERS_RAW, _CATALOG, server._estimate_listing_economics)
    wb = load_workbook(buf)
    ws = wb["Orders"]
    row = [c.value for c in ws[5]]
    check(row[0] == 999, f"expected receipt_id 999, got: {row}")
    check(row[2] == "Jane Buyer", f"expected buyer name, got: {row}")
    check(row[4] == 14.99, f"expected total 14.99 from grandtotal, got: {row}")


def test_blank_template_sheets_carry_no_fake_live_data():
    # Physical Inventory etc. have no live data source in this codebase --
    # they must stay honest empty scaffolds (label rows present, quantity/cost
    # columns blank), not a false "live" claim.
    buf = business_tracker.build_workbook(_LISTINGS, _SALES, _ORDERS_RAW, _CATALOG, server._estimate_listing_economics)
    wb = load_workbook(buf)
    ws = wb["Physical Inventory"]
    row = [c.value for c in ws[5]]  # first material row: ("Filament spool", "PLA / PLA+", "", "", "", "")
    check(row[0] == "Filament spool", f"expected a labeled row, got: {row}")
    check(row[4] in (None, ""), f"Qty on Hand must be blank (no live source), got: {row[4]!r}")


def test_empty_inputs_do_not_crash():
    buf = business_tracker.build_workbook([], {}, [], [], server._estimate_listing_economics)
    wb = load_workbook(buf)
    check("Dashboard" in wb.sheetnames, "an empty-data workbook must still build successfully")


# ── 2. orders-fetch dedup ───────────────────────────────────────────────────
def test_get_recent_orders_raw_is_the_single_fetch_point():
    calls = {"n": 0}

    class _FakeClient:
        def get_orders(self, limit=100):
            calls["n"] += 1
            return {"results": _ORDERS_RAW}

    orig_client_cls = server.EtsyAPIClient
    orig_cache = dict(server._cache)
    try:
        server.EtsyAPIClient = _FakeClient
        server._cache.clear()
        raw1 = server._get_recent_orders_raw()
        raw2 = server._sales_by_listing_sync()  # must reuse the cache, not re-fetch
        raw3 = server._search_orders("")  # must also reuse the cache
        check(calls["n"] == 1, f"expected exactly 1 real Etsy call across all 3 callers (shared cache), got {calls['n']}")
        check(raw1 == _ORDERS_RAW, f"unexpected raw orders: {raw1}")
        check(raw2.get(111) == 1, f"sales map should count the fixture's 1 unit for listing 111, got: {raw2}")
        check(len(raw3) == 1, f"search with empty query should match the 1 fixture order, got: {raw3}")
    finally:
        server.EtsyAPIClient = orig_client_cls
        server._cache.clear()
        server._cache.update(orig_cache)


def test_get_recent_orders_raw_degrades_to_empty_list_on_failure():
    class _FailingClient:
        def get_orders(self, limit=100):
            raise RuntimeError("simulated Etsy outage")

    orig_client_cls = server.EtsyAPIClient
    orig_cache = dict(server._cache)
    try:
        server.EtsyAPIClient = _FailingClient
        server._cache.clear()
        result = server._get_recent_orders_raw()
        check(result == [], f"a failed fetch must degrade to an empty list, not raise, got: {result}")
    finally:
        server.EtsyAPIClient = orig_client_cls
        server._cache.clear()
        server._cache.update(orig_cache)


# ── 3. GET /api/business-tracker.xlsx endpoint ──────────────────────────────
def test_endpoint_requires_auth():
    c = TestClient(server.app, base_url="https://testserver")
    resp = c.get("/api/business-tracker.xlsx")
    check(resp.status_code == 401, f"unauthenticated request should 401, got {resp.status_code}")


def test_endpoint_returns_a_real_downloadable_workbook():
    orig_listings = server._listings_sync
    orig_orders = server._get_recent_orders_raw
    orig_sales = server._sales_by_listing_sync
    orig_catalog_read = None
    try:
        server._listings_sync = lambda state="active": {"listings": _LISTINGS, "count": 1, "state": "active"}
        server._get_recent_orders_raw = lambda: _ORDERS_RAW
        server._sales_by_listing_sync = lambda: _SALES

        c = TestClient(server.app, base_url="https://testserver")
        resp = c.get("/api/business-tracker.xlsx", headers={"Authorization": f"Bearer {os.environ['APP_SECRET_TOKEN']}"})
        check(resp.status_code == 200, f"expected 200, got {resp.status_code}: {resp.text[:300]}")
        check(
            resp.headers.get("content-type", "") ==
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            f"unexpected content-type: {resp.headers.get('content-type')}",
        )
        check("attachment" in resp.headers.get("content-disposition", ""),
              f"expected an attachment disposition, got: {resp.headers.get('content-disposition')}")
        check(".xlsx" in resp.headers.get("content-disposition", ""),
              f"expected .xlsx filename, got: {resp.headers.get('content-disposition')}")

        wb = load_workbook(io.BytesIO(resp.content))
        check("COGS & Profit" in wb.sheetnames, f"downloaded workbook missing expected sheet, got: {wb.sheetnames}")
    finally:
        server._listings_sync = orig_listings
        server._get_recent_orders_raw = orig_orders
        server._sales_by_listing_sync = orig_sales


def test_endpoint_degrades_gracefully_when_etsy_is_unreachable():
    # Best-effort per-lookup tolerance: a listings/orders fetch failure must
    # still return a downloadable workbook (with those sheets empty), not a
    # 500 -- same tolerance _sales_by_listing_sync()/_compute_cogs_status()
    # already apply.
    orig_listings = server._listings_sync

    def _boom(state="active"):
        raise RuntimeError("simulated Etsy outage")

    try:
        server._listings_sync = _boom
        c = TestClient(server.app, base_url="https://testserver")
        resp = c.get("/api/business-tracker.xlsx", headers={"Authorization": f"Bearer {os.environ['APP_SECRET_TOKEN']}"})
        check(resp.status_code == 200, f"a listings-fetch failure must still produce a downloadable workbook, got {resp.status_code}")
        wb = load_workbook(io.BytesIO(resp.content))
        check("COGS & Profit" in wb.sheetnames, "the workbook must still build with an empty COGS & Profit sheet")
    finally:
        server._listings_sync = orig_listings


# ── 4. Create-screen overlay entries merged into the Products sheet (2026-07-25) ──
# Scott: "document it in the excel file" -- a product built via the Create
# screen's "+ new one" flow only ever exists in product_catalog_overrides.json,
# never in the git-tracked data/product_catalog.json this endpoint reads
# directly. Without the merge in get_business_tracker()'s _gather_and_build(),
# it silently never showed up in the downloaded workbook.

def test_products_sheet_includes_new_product_overlay_entries():
    orig_overrides = server._product_catalog_overrides
    orig_listings = server._listings_sync
    orig_orders = server._get_recent_orders_raw
    orig_sales = server._sales_by_listing_sync
    try:
        server._product_catalog_overrides = lambda: {
            "COLOR9999": {
                "is_new_product": True, "product_id": "COLOR9999",
                "name": "ocean animals", "category": "coloring_pages",
                "status": "draft", "price": None, "etsy_listing_id": "",
                "created_at": "2026-07-25T00:00:00+00:00",
            },
        }
        server._listings_sync = lambda state="active": {"listings": _LISTINGS, "count": 1, "state": "active"}
        server._get_recent_orders_raw = lambda: _ORDERS_RAW
        server._sales_by_listing_sync = lambda: _SALES

        c = TestClient(server.app, base_url="https://testserver")
        resp = c.get("/api/business-tracker.xlsx", headers={"Authorization": f"Bearer {os.environ['APP_SECRET_TOKEN']}"})
        check(resp.status_code == 200, f"expected 200, got {resp.status_code}: {resp.text[:300]}")
        wb = load_workbook(io.BytesIO(resp.content))
        ws = wb["Products"]
        rows = [[c.value for c in row] for row in ws.iter_rows(min_row=5)]
        matches = [r for r in rows if r[0] == "COLOR9999"]
        check(len(matches) == 1, f"the overlay-only product must appear exactly once in Products, got {matches}")
        if matches:
            check(matches[0][2] == "coloring_pages", f"got row {matches[0]}")
            check(matches[0][3] == "draft", f"got row {matches[0]}")
    finally:
        server._product_catalog_overrides = orig_overrides
        server._listings_sync = orig_listings
        server._get_recent_orders_raw = orig_orders
        server._sales_by_listing_sync = orig_sales


def test_products_sheet_does_not_duplicate_overlay_entry_already_in_base_catalog():
    """An override entry whose pid ALREADY exists in the real base catalog
    (e.g. a status patch on a published product, not a Create-screen new
    product) must never be synthesized into a second row."""
    orig_overrides = server._product_catalog_overrides
    orig_catalog_read = Path("data/product_catalog.json").read_text
    try:
        real_catalog = json.loads(orig_catalog_read())
        known_pid = real_catalog[0]["product_id"] if real_catalog else "DP1026"
        server._product_catalog_overrides = lambda: {
            known_pid: {"is_new_product": True, "product_id": known_pid, "etsy_listing_id": "999"},
        }
        c = TestClient(server.app, base_url="https://testserver")
        resp = c.get("/api/business-tracker.xlsx", headers={"Authorization": f"Bearer {os.environ['APP_SECRET_TOKEN']}"})
        check(resp.status_code == 200, f"expected 200, got {resp.status_code}")
        wb = load_workbook(io.BytesIO(resp.content))
        ws = wb["Products"]
        rows = [[c.value for c in row] for row in ws.iter_rows(min_row=5)]
        matches = [r for r in rows if r[0] == known_pid]
        check(len(matches) == 1, f"an override for an existing base-catalog pid must not be duplicated, got {matches}")
    finally:
        server._product_catalog_overrides = orig_overrides


def run() -> None:
    for fn in [v for k, v in sorted(globals().items()) if k.startswith("test_")]:
        try:
            fn()
        except Exception:  # noqa: BLE001
            _failures.append(f"{fn.__name__} raised:\n{traceback.format_exc()}")
    if _failures:
        print("BUSINESS TRACKER TESTS FAILED:")
        for f in _failures:
            print(" -", f)
        sys.exit(1)
    print("BUSINESS TRACKER TESTS OK — workbook builds with all 9 sheets from live-shaped "
          "fixture data, Products/COGS & Profit/Orders reflect the real inputs, blank "
          "template sheets stay honest, the orders-fetch dedup is a single real call "
          "shared by all 3 callers, and the endpoint requires auth, returns a real "
          "downloadable .xlsx, and degrades gracefully on an Etsy fetch failure.")


if __name__ == "__main__":
    run()
