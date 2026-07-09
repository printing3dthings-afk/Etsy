#!/usr/bin/env python3
"""
Generates OnBrandCraftz_Business_Tracker.xlsx -- a working spreadsheet for Scott to
track products, physical inventory/consumables, suppliers, per-SKU cost/margin,
equipment, and tax-deductible expenses. Pre-filled with what's already known from
CLAUDE.md's product catalog and business docs rather than a blank template; Scott
fills in real counts/costs/suppliers only he knows.

This is a one-off deliverable generator (not a runtime app dependency) -- openpyxl
isn't added to requirements.txt for that reason, same as PyInstaller isn't.

Run:  python tools/generate_business_tracker.py
Output: written next to this script's invocation dir (see OUT_PATH below).
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

OUT_PATH = Path("OnBrandCraftz_Business_Tracker.xlsx")

# ── shared styling (matches the app's cyan/navy palette, kept professional) ────────
HEADER_FILL = PatternFill("solid", fgColor="1B2568")  # midnight blue, same family as DP1028
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14, color="1B2568")
SUBTLE_FONT = Font(italic=True, size=9, color="6B7280")
THIN_BORDER = Border(*(Side(style="thin", color="D1D5DB") for _ in range(4)))
LOW_STOCK_FILL = PatternFill("solid", fgColor="FDE2E2")
LOW_STOCK_FONT = Font(color="9B1C1C")


def _style_header_row(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 32


def _autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _title_block(ws, title, subtitle, ncols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws.cell(row=1, column=1, value=title).font = TITLE_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    ws.cell(row=2, column=1, value=subtitle).font = SUBTLE_FONT
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 16


def _write_table(ws, headers, rows, start_row, widths):
    _style_header_row(ws, start_row, len(headers))
    for c, h in enumerate(headers, start=1):
        ws.cell(row=start_row, column=c, value=h)
    r = start_row + 1
    for row in rows:
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = THIN_BORDER
        r += 1
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)
    _autosize(ws, widths)
    return r  # next empty row


# ── Sheet 1: Products ───────────────────────────────────────────────────────────
def build_products(wb):
    ws = wb.active
    ws.title = "Products"
    _title_block(ws, "Products", "Master catalog -- every SKU, digital and physical.", 9)
    headers = ["SKU", "Name", "Type", "Etsy Listing ID", "Status", "Price",
               "Color Theme", "Launch Date", "Notes"]
    rows = [
        ("DP1026", "Ultimate Digital Life Planner", "Digital Planner", "", "Live", 14.99,
         "Lavender Dreams", "", "143 pages + sticker pack (11 sheets/328 stickers)"),
        ("DP1027", "Student & School Planner 2026", "Digital Planner", "", "Live", 9.99,
         "Cotton Candy", "", "131 pages + sticker pack (11 sheets/320 stickers)"),
        ("DP1028", "Budget & Finance Planner 2026", "Digital Planner", "", "Live", 12.99,
         "Midnight Blue", "", "144 pages + sticker pack (11 sheets/419 stickers)"),
        ("DP1029", "Fitness & Wellness Planner 2026", "Digital Planner", "", "Live", 12.99,
         "Coral Peach", "", "133 pages + sticker pack (11 sheets/377 stickers)"),
        ("DP1030", "(expanded catalog -- fill in title)", "Digital Planner", "", "Files exist, not yet listed", "",
         "", "", "Files on disk in data/digital_products/ -- documentation pending"),
        ("DP1031", "(expanded catalog -- fill in title)", "Digital Planner", "", "Files exist, not yet listed", "",
         "", "", "Files on disk in data/digital_products/ -- documentation pending"),
        ("DP1032", "(expanded catalog -- fill in title)", "Digital Planner", "", "Files exist, not yet listed", "",
         "", "", "Files on disk in data/digital_products/ -- documentation pending"),
        ("DP1033", "(expanded catalog -- fill in title)", "Digital Planner", "", "Files exist, not yet listed", "",
         "", "", "Files on disk in data/digital_products/ -- documentation pending"),
        ("DP1034", "(expanded catalog -- fill in title)", "Digital Planner", "", "Files exist, not yet listed", "",
         "", "", "Files on disk in data/digital_products/ -- documentation pending"),
        ("SS-series", "3D Print SVG Packs (multiple SKUs)", "SVG Pack / 3D Print", "", "See data/digital_products/",
         "", "", "", "Add one row per SS#### SKU as they're catalogued"),
        ("", "", "", "", "", "", "", "", ""),
        ("", "", "", "", "", "", "", "", ""),
        ("", "", "", "", "", "", "", "", ""),
    ]
    _write_table(ws, headers, rows, 4, [10, 32, 16, 14, 22, 9, 16, 12, 42])


# ── Sheet 2: Physical Inventory ─────────────────────────────────────────────────
def build_physical_inventory(wb):
    ws = wb.create_sheet("Physical Inventory")
    _title_block(ws, "Physical Inventory", "Filament, nozzles, build plates -- Bambu Lab P1S consumables.", 6)
    headers = ["Item", "Material/Type", "Color", "Brand", "Qty on Hand", "Unit Cost"]
    materials = ["PLA / PLA+", "Silk PLA", "High-Speed PLA", "PETG", "PETG-CF", "TPU",
                 "ABS", "ASA", "PA / Nylon", "PA-CF / PLA-CF", "PC", "PVA"]
    rows = [("Filament spool", m, "", "", "", "") for m in materials]
    rows += [
        ("Nozzle", "Brass 0.4mm (standard)", "", "", "", ""),
        ("Nozzle", "Hardened steel 0.4mm (CF/GF filaments)", "", "", "", ""),
        ("Nozzle", "0.2mm (fine detail)", "", "", "", ""),
        ("Nozzle", "0.6mm (fast/low detail)", "", "", "", ""),
        ("Nozzle", "0.8mm (fast/low detail)", "", "", "", ""),
        ("Build plate", "Textured PEI", "", "", "", ""),
        ("Build plate", "Smooth PEI", "", "", "", ""),
        ("", "", "", "", "", ""),
        ("", "", "", "", "", ""),
    ]
    _write_table(ws, headers, rows, 4, [16, 30, 14, 16, 12, 12])


# ── Sheet 3: Consumables & Reorder ──────────────────────────────────────────────
def build_consumables(wb):
    ws = wb.create_sheet("Consumables & Reorder")
    _title_block(ws, "Consumables & Reorder",
                 "Everything that runs out -- adhesives, packaging, print consumables. "
                 "Row highlights red when Current Stock <= Reorder Point.", 8)
    headers = ["Item", "Category", "Current Stock", "Reorder Point", "Reorder Qty",
               "Supplier", "Lead Time (days)", "Last Ordered"]
    rows = [
        ("Glue stick / bed adhesive", "Print consumable", "", "", "", "", "", ""),
        ("Isopropyl alcohol", "Print consumable", "", "", "", "", "", ""),
        ("Nozzle wipe / cleaning filament", "Print consumable", "", "", "", "", "", ""),
        ("Poly mailers", "Packaging", "", "", "", "", "", ""),
        ("Shipping labels", "Packaging", "", "", "", "", "", ""),
        ("Tissue paper / branded inserts", "Packaging", "", "", "", "", "", ""),
        ("Desiccant packs (filament storage)", "Storage", "", "", "", "", "", ""),
        ("", "", "", "", "", "", "", ""),
        ("", "", "", "", "", "", "", ""),
        ("", "", "", "", "", "", "", ""),
    ]
    next_row = _write_table(ws, headers, rows, 4, [30, 16, 14, 14, 12, 18, 14, 14])
    # Highlight any data row where Current Stock <= Reorder Point (both numeric, C/D cols)
    data_range = f"C5:C{next_row - 1}"
    ws.conditional_formatting.add(
        data_range,
        CellIsRule(operator="lessThanOrEqual", formula=["$D5"], fill=LOW_STOCK_FILL, font=LOW_STOCK_FONT),
    )


# ── Sheet 4: Suppliers ───────────────────────────────────────────────────────────
def build_suppliers(wb):
    ws = wb.create_sheet("Suppliers")
    _title_block(ws, "Suppliers", "Where everything comes from.", 5)
    headers = ["Name", "What They Supply", "Contact / Account", "Lead Time", "Notes"]
    rows = [("", "", "", "", "") for _ in range(10)]
    _write_table(ws, headers, rows, 4, [22, 28, 26, 14, 34])


# ── Sheet 5: COGS & Pricing ──────────────────────────────────────────────────────
def build_cogs_pricing(wb):
    ws = wb.create_sheet("COGS & Pricing")
    _title_block(ws, "COGS & Pricing",
                 "Etsy fees calculated live from CLAUDE.md's documented fee structure: "
                 "6.5% transaction + $0.20 listing + 3%+$0.25 processing.", 7)
    headers = ["SKU", "Etsy Price", "Etsy Fees (calc)", "Material/File Cost",
               "Net Profit (calc)", "Margin % (calc)", "Notes"]
    known = [
        ("DP1026", 14.99),
        ("DP1027", 9.99),
        ("DP1028", 12.99),
        ("DP1029", 12.99),
    ]
    start_row = 4
    _style_header_row(ws, start_row, len(headers))
    for c, h in enumerate(headers, start=1):
        ws.cell(row=start_row, column=c, value=h)
    r = start_row + 1
    for sku, price in known:
        ws.cell(row=r, column=1, value=sku).border = THIN_BORDER
        ws.cell(row=r, column=2, value=price).border = THIN_BORDER
        # Etsy fees: 6.5% transaction + $0.20 listing + (3% + $0.25) processing
        ws.cell(row=r, column=3,
                value=f"=ROUND(B{r}*0.065 + 0.20 + B{r}*0.03 + 0.25, 2)").border = THIN_BORDER
        ws.cell(row=r, column=4, value=0).border = THIN_BORDER  # digital = ~$0 marginal cost
        ws.cell(row=r, column=5, value=f"=B{r}-C{r}-D{r}").border = THIN_BORDER
        ws.cell(row=r, column=6, value=f"=ROUND(E{r}/B{r}*100,1)").border = THIN_BORDER
        ws.cell(row=r, column=7, value="Digital -- material cost is ~$0 (no COGS beyond time)").border = THIN_BORDER
        r += 1
    for _ in range(8):
        for c in range(1, 8):
            ws.cell(row=r, column=c).border = THIN_BORDER
        r += 1
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)
    _autosize(ws, [10, 12, 14, 16, 14, 12, 40])


# ── Sheet 6: Equipment & Assets ──────────────────────────────────────────────────
def build_equipment(wb):
    ws = wb.create_sheet("Equipment & Assets")
    _title_block(ws, "Equipment & Assets",
                 "Section 179 / bonus depreciation candidates -- see CLAUDE.md's "
                 "Business Structure & Tax section.", 6)
    headers = ["Item", "Purchase Date", "Cost", "Depreciation Note", "Maintenance Log", "Notes"]
    rows = [
        ("Bambu Lab P1S", "", "", "Section 179 (2026 limit: $2.56M) or 100% bonus depreciation", "", ""),
        ("AMS 2 Pro", "", "", "Section 179 or bonus depreciation", "", ""),
        ("", "", "", "", "", ""),
        ("", "", "", "", "", ""),
        ("", "", "", "", "", ""),
    ]
    _write_table(ws, headers, rows, 4, [22, 14, 12, 40, 26, 26])


# ── Sheet 7: Expense & Tax Tracker ──────────────────────────────────────────────
def build_expense_tax(wb):
    ws = wb.create_sheet("Expense & Tax Tracker")
    _title_block(ws, "Expense & Tax Tracker",
                 "Mirrors CLAUDE.md's Key Deductions table -- log as you go through the year.", 6)
    headers = ["Date", "Category", "Amount", "Deductible?", "Schedule C Line", "Notes"]
    categories = [
        ("Filament / materials (COGS)", "Y", "Part III, Line 36"),
        ("Equipment (Section 179 / depreciation)", "Y", "Line 13"),
        ("Canva Pro / Adobe / AI subscriptions (Claude, OpenAI)", "Y", "Line 27a"),
        ("Etsy fees (6.5% + 3%+$0.25)", "Y", "Line 10"),
        ("Home studio (dedicated space)", "Y", "Form 8829 or $5/sq ft simplified"),
        ("Quality print samples photographed", "Y", "Line 22 (advertising) -- NOT COGS"),
        ("Vehicle mileage (post office runs)", "Y", "$0.725/mile (2026 rate)"),
    ]
    rows = [("", cat, "", ded, line, "") for cat, ded, line in categories]
    rows += [("", "", "", "", "", "") for _ in range(10)]
    _write_table(ws, headers, rows, 4, [12, 42, 12, 14, 30, 26])


# ── Sheet 8: Dashboard ───────────────────────────────────────────────────────────
def build_dashboard(wb, sheet_order):
    ws = wb.create_sheet("Dashboard", 0)  # placed first
    _title_block(ws, "OnBrandCraftz Business Tracker", "Summary -- fill in the other sheets, this rolls up live.", 2)
    ws.cell(row=4, column=1, value="Metric").font = HEADER_FONT
    ws.cell(row=4, column=1).fill = HEADER_FILL
    ws.cell(row=4, column=2, value="Value").font = HEADER_FONT
    ws.cell(row=4, column=2).fill = HEADER_FILL
    metrics = [
        ("Total products listed", '=COUNTIF(Products!E5:E20,"Live")'),
        ("Products pending documentation", '=COUNTIF(Products!E5:E20,"Files exist, not yet listed")'),
        ("Consumables at/below reorder point",
         '=SUMPRODUCT((\'Consumables & Reorder\'!C5:C14<>"")*(\'Consumables & Reorder\'!C5:C14<=\'Consumables & Reorder\'!D5:D14))'),
        ("Average margin % (COGS & Pricing sheet)", "=ROUND(AVERAGE('COGS & Pricing'!F5:F8),1)"),
    ]
    r = 5
    for label, formula in metrics:
        ws.cell(row=r, column=1, value=label).border = THIN_BORDER
        ws.cell(row=r, column=2, value=formula).border = THIN_BORDER
        r += 1
    _autosize(ws, [40, 16])
    ws.sheet_view.showGridLines = False


def main():
    wb = Workbook()
    build_products(wb)
    build_physical_inventory(wb)
    build_consumables(wb)
    build_suppliers(wb)
    build_cogs_pricing(wb)
    build_equipment(wb)
    build_expense_tax(wb)
    build_dashboard(wb, None)
    wb.save(OUT_PATH)
    print(f"Wrote {OUT_PATH.resolve()}")
    print("Sheets:", wb.sheetnames)


if __name__ == "__main__":
    main()
