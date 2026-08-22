#!/usr/bin/env python3
"""
Builds the OnBrandCraftz Business Tracker workbook -- one .xlsx, multiple
tabs, generated fresh from live data on every download (see
GET /api/business-tracker.xlsx in tools/api_server/main.py).

Adapted 2026-07-18 from the archived one-off tools/generate_business_tracker.py
(data/trash/files/20260711-048__generate_business_tracker.py -- see
data/trash/DELETED.md for the removal record). That script hardcoded a
2026-05 product snapshot and wrote to the gitignored data/backups/ folder,
so it never actually reached the hosted Frank Files screen. This version is
a pure builder (listings/catalog/orders/sales passed in, no Etsy/filesystem
calls of its own) so it can be called directly from a FastAPI endpoint and
unit-tested with fixture data -- the styling helpers below are carried over
unchanged from the archived script (they were already generic).

Sheets:
  Dashboard            -- rollup formulas over the other sheets
  Products              -- from data/product_catalog.json (all products,
                            including drafts not yet live on Etsy)
  COGS & Profit          -- from live active Etsy listings + the same
                            fee/COGS math powering the /api/cogs-status HUD
                            card (main.py's _estimate_listing_economics)
  Orders                 -- live, last 100 paid receipts
  Physical Inventory / Consumables & Reorder / Suppliers / Equipment &
  Assets / Expense & Tax Tracker -- blank labeled templates for Scott to
                            fill in by hand; no live data source exists for
                            these in this codebase, so they stay honest
                            empty scaffolds rather than fake "live" tabs.

Not a runtime app dependency in the PyInstaller sense, but openpyxl IS a
real pinned requirement now (requirements.txt) since this is wired to a
live endpoint, unlike the archived script's one-off status.
"""
from __future__ import annotations

import io
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── shared styling (matches the app's cyan/navy palette) ───────────────────
HEADER_FILL = PatternFill("solid", fgColor="1B2568")  # midnight blue, DP1028 family
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14, color="1B2568")
SUBTLE_FONT = Font(italic=True, size=9, color="6B7280")
THIN_BORDER = Border(*(Side(style="thin", color="D1D5DB") for _ in range(4)))
LOW_STOCK_FILL = PatternFill("solid", fgColor="FDE2E2")
LOW_STOCK_FONT = Font(color="9B1C1C")


def _style_header_row(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 32


def _autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _title_block(ws, title, subtitle, ncols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws.cell(row=1, column=1, value=title).font = TITLE_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    ws.cell(row=2, column=1, value=subtitle).font = SUBTLE_FONT
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 16


def _write_table(ws, headers, rows, start_row, widths):
    _style_header_row(ws, start_row, len(headers))
    for c, h in enumerate(headers, start=1):
        ws.cell(row=start_row, column=c, value=h)
    r = start_row + 1
    for row in rows:
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = THIN_BORDER
        r += 1
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)
    _autosize(ws, widths)
    return r  # next empty row


def _receipt_total(receipt: dict) -> float:
    gt = receipt.get("grandtotal", {}) or {}
    divisor = gt.get("divisor", 100) or 100
    return round((gt.get("amount", 0) or 0) / divisor, 2)


def _receipt_date(receipt: dict) -> str:
    ts = receipt.get("create_timestamp")
    if not ts:
        return ""
    try:
        return datetime.fromtimestamp(int(ts), tz=timezone.utc).strftime("%Y-%m-%d")
    except (TypeError, ValueError, OSError):
        return ""


# ── Sheet: Products (live, from data/product_catalog.json) ────────────────
def _build_products(wb, catalog: list[dict]):
    ws = wb.active
    ws.title = "Products"
    _title_block(ws, "Products", "Full catalog -- every product, live or still in progress.", 8)
    headers = ["Product ID", "Name", "Category", "Status", "Price",
               "Etsy Listing ID", "Last Updated", "Notes"]
    rows = [
        (
            p.get("product_id", ""),
            p.get("name", ""),
            p.get("category", ""),
            p.get("status", ""),
            p.get("price", ""),
            p.get("etsy_listing_id", ""),
            p.get("last_updated", ""),
            (p.get("note") or "")[:200],
        )
        for p in catalog
    ]
    _write_table(ws, headers, rows, 4, [12, 36, 16, 16, 9, 16, 14, 50])


# ── Sheet: COGS & Profit (live, from active Etsy listings) ─────────────────
def _build_cogs_profit(wb, listings: list[dict], sales: dict, econ_fn):
    ws = wb.create_sheet("COGS & Profit")
    _title_block(
        ws, "COGS & Profit",
        "Real Etsy price + real documented fees (6.5% transaction + 3%+$0.25 "
        "processing + $0.20 listing) minus an ESTIMATED COGS (digital $0, "
        "3D-print physical $7.50/unit typical). Recent units sold are real "
        "(last 100 paid receipts). Not real accounting -- an estimate.",
        9,
    )
    headers = ["Listing ID", "Title", "Price", "Product Type", "Recent Units Sold",
               "Fees (real)", "COGS (est.)", "Net Profit (est.)", "Margin %"]
    rows = []
    for l in listings:
        price = l.get("price") or 0
        econ = econ_fn(price, l.get("title", ""))
        units = sales.get(l.get("listing_id"), 0)
        rows.append((
            l.get("listing_id"),
            (l.get("title") or "")[:60],
            price,
            econ["product_type"],
            units,
            econ["fees_real"],
            econ["cogs_estimate"],
            round(econ["net_estimate"] * units, 2),
            econ["margin_pct"],
        ))
    _write_table(ws, headers, rows, 4, [14, 42, 9, 14, 16, 12, 12, 16, 10])


# ── Sheet: Orders (live, last 100 paid receipts) ────────────────────────────
def _build_orders(wb, orders_raw: list[dict]):
    ws = wb.create_sheet("Orders")
    _title_block(ws, "Orders", "Live -- last 100 paid receipts, pulled fresh from Etsy on every download.", 5)
    headers = ["Order ID", "Date", "Buyer", "Items", "Total"]
    rows = []
    for r in orders_raw:
        item_titles = ", ".join((t.get("title") or "")[:40] for t in (r.get("transactions") or []))
        rows.append((
            r.get("receipt_id", ""),
            _receipt_date(r),
            r.get("name", "") or "",
            item_titles,
            _receipt_total(r),
        ))
    _write_table(ws, headers, rows, 4, [14, 12, 22, 60, 10])


# ── Blank labeled templates (no live data source exists for these) ─────────
def _build_physical_inventory(wb):
    ws = wb.create_sheet("Physical Inventory")
    _title_block(ws, "Physical Inventory", "Filament, nozzles, build plates -- Bambu Lab P1S consumables.", 6)
    headers = ["Item", "Material/Type", "Color", "Brand", "Qty on Hand", "Unit Cost"]
    materials = ["PLA / PLA+", "Silk PLA", "High-Speed PLA", "PETG", "PETG-CF", "TPU",
                 "ABS", "ASA", "PA / Nylon", "PA-CF / PLA-CF", "PC", "PVA"]
    rows = [("Filament spool", m, "", "", "", "") for m in materials]
    rows += [
        ("Nozzle", "Brass 0.4mm (standard)", "", "", "", ""),
        ("Nozzle", "Hardened steel 0.4mm (CF/GF filaments)", "", "", "", ""),
        ("Nozzle", "0.2mm (fine detail)", "", "", "", ""),
        ("Nozzle", "0.6mm (fast/low detail)", "", "", "", ""),
        ("Nozzle", "0.8mm (fast/low detail)", "", "", "", ""),
        ("Build plate", "Textured PEI", "", "", "", ""),
        ("Build plate", "Smooth PEI", "", "", "", ""),
        ("", "", "", "", "", ""),
        ("", "", "", "", "", ""),
    ]
    _write_table(ws, headers, rows, 4, [16, 30, 14, 16, 12, 12])


def _build_consumables(wb):
    ws = wb.create_sheet("Consumables & Reorder")
    _title_block(ws, "Consumables & Reorder",
                 "Everything that runs out -- adhesives, packaging, print consumables. "
                 "Row highlights red when Current Stock <= Reorder Point.", 8)
    headers = ["Item", "Category", "Current Stock", "Reorder Point", "Reorder Qty",
               "Supplier", "Lead Time (days)", "Last Ordered"]
    rows = [
        ("Glue stick / bed adhesive", "Print consumable", "", "", "", "", "", ""),
        ("Isopropyl alcohol", "Print consumable", "", "", "", "", "", ""),
        ("Nozzle wipe / cleaning filament", "Print consumable", "", "", "", "", "", ""),
        ("Poly mailers", "Packaging", "", "", "", "", "", ""),
        ("Shipping labels", "Packaging", "", "", "", "", "", ""),
        ("Tissue paper / branded inserts", "Packaging", "", "", "", "", "", ""),
        ("Desiccant packs (filament storage)", "Storage", "", "", "", "", "", ""),
        ("", "", "", "", "", "", "", ""),
        ("", "", "", "", "", "", "", ""),
        ("", "", "", "", "", "", "", ""),
    ]
    next_row = _write_table(ws, headers, rows, 4, [30, 16, 14, 14, 12, 18, 14, 14])
    data_range = f"C5:C{next_row - 1}"
    ws.conditional_formatting.add(
        data_range,
        CellIsRule(operator="lessThanOrEqual", formula=["$D5"], fill=LOW_STOCK_FILL, font=LOW_STOCK_FONT),
    )


def _build_suppliers(wb):
    ws = wb.create_sheet("Suppliers")
    _title_block(ws, "Suppliers", "Where everything comes from.", 5)
    headers = ["Name", "What They Supply", "Contact / Account", "Lead Time", "Notes"]
    rows = [("", "", "", "", "") for _ in range(10)]
    _write_table(ws, headers, rows, 4, [22, 28, 26, 14, 34])


def _build_equipment(wb):
    ws = wb.create_sheet("Equipment & Assets")
    _title_block(ws, "Equipment & Assets",
                 "Section 179 / bonus depreciation candidates -- see CLAUDE.md's "
                 "Business Structure & Tax section.", 6)
    headers = ["Item", "Purchase Date", "Cost", "Depreciation Note", "Maintenance Log", "Notes"]
    rows = [
        ("Bambu Lab P1S", "", "", "Section 179 (2026 limit: $2.56M) or 100% bonus depreciation", "", ""),
        ("AMS 2 Pro", "", "", "Section 179 or bonus depreciation", "", ""),
        ("", "", "", "", "", ""),
        ("", "", "", "", "", ""),
        ("", "", "", "", "", ""),
    ]
    _write_table(ws, headers, rows, 4, [22, 14, 12, 40, 26, 26])


def _build_expense_tax(wb):
    ws = wb.create_sheet("Expense & Tax Tracker")
    _title_block(ws, "Expense & Tax Tracker",
                 "Mirrors CLAUDE.md's Key Deductions table -- log as you go through the year.", 6)
    headers = ["Date", "Category", "Amount", "Deductible?", "Schedule C Line", "Notes"]
    categories = [
        ("Filament / materials (COGS)", "Y", "Part III, Line 36"),
        ("Equipment (Section 179 / depreciation)", "Y", "Line 13"),
        ("Canva Pro / Adobe / AI subscriptions (Claude, OpenAI)", "Y", "Line 27a"),
        ("Etsy fees (6.5% + 3%+$0.25)", "Y", "Line 10"),
        ("Home studio (dedicated space)", "Y", "Form 8829 or $5/sq ft simplified"),
        ("Quality print samples photographed", "Y", "Line 22 (advertising) -- NOT COGS"),
        ("Vehicle mileage (post office runs)", "Y", "$0.725/mile (2026 rate)"),
    ]
    rows = [("", cat, "", ded, line, "") for cat, ded, line in categories]
    rows += [("", "", "", "", "", "") for _ in range(10)]
    _write_table(ws, headers, rows, 4, [12, 42, 12, 14, 30, 26])


# ── Sheet: Dashboard (first tab, rollup formulas) ───────────────────────────
def _build_dashboard(wb, catalog_count: int, listing_count: int):
    ws = wb.create_sheet("Dashboard", 0)
    _title_block(ws, "OnBrandCraftz Business Tracker",
                 "Generated fresh from live data on every download.", 2)
    ws.cell(row=4, column=1, value="Metric").font = HEADER_FONT
    ws.cell(row=4, column=1).fill = HEADER_FILL
    ws.cell(row=4, column=2, value="Value").font = HEADER_FONT
    ws.cell(row=4, column=2).fill = HEADER_FILL
    metrics = [
        ("Total products in catalog", catalog_count),
        ("Active Etsy listings", listing_count),
        ("Consumables at/below reorder point",
         '=SUMPRODUCT((\'Consumables & Reorder\'!C5:C14<>"")*(\'Consumables & Reorder\'!C5:C14<=\'Consumables & Reorder\'!D5:D14))'),
        ("Average margin % (COGS & Profit sheet)",
         "=IFERROR(ROUND(AVERAGE('COGS & Profit'!I5:I500),1),\"\")"),
        ("Total recent units sold (COGS & Profit sheet)",
         "=SUM('COGS & Profit'!E5:E500)"),
    ]
    r = 5
    for label, val in metrics:
        ws.cell(row=r, column=1, value=label).border = THIN_BORDER
        ws.cell(row=r, column=2, value=val).border = THIN_BORDER
        r += 1
    _autosize(ws, [42, 16])
    ws.sheet_view.showGridLines = False


def build_workbook(
    listings: list[dict],
    sales: dict,
    orders_raw: list[dict],
    catalog: list[dict],
    econ_fn,
) -> io.BytesIO:
    """Builds the full multi-tab workbook and returns it as an in-memory
    BytesIO (never written to disk -- streamed straight back by the FastAPI
    endpoint). `econ_fn` is main.py's _estimate_listing_economics, passed in
    rather than imported, so this module has no dependency on main.py and
    stays independently testable with fixture data."""
    wb = Workbook()
    _build_products(wb, catalog)
    _build_cogs_profit(wb, listings, sales, econ_fn)
    _build_orders(wb, orders_raw)
    _build_physical_inventory(wb)
    _build_consumables(wb)
    _build_suppliers(wb)
    _build_equipment(wb)
    _build_expense_tax(wb)
    _build_dashboard(wb, len(catalog), len(listings))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
