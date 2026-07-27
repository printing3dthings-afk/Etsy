#!/usr/bin/env python3
"""
OnBrandCraftz — SEO Title & Tag Optimizer
==========================================
Audits all active Etsy listing titles against 2026 best practices and
generates optimized alternatives with buyer-intent keywords.

Rules applied:
  1. First 40 characters must contain the primary buyer-intent keyword
  2. Use pipe separators (|) for readability
  3. Include room/style/mood descriptors (Boho, Nursery, Kitchen, etc.)
  4. All 13 tag slots should be filled with unique multi-word phrases
  5. No single-word tags (Etsy penalizes these)
  6. No tag over 20 characters
  7. No duplicate words across title and tags unnecessarily

Usage:
  python tools/seo_title_optimizer.py              # audit all active listings
  python tools/seo_title_optimizer.py --fix         # show suggested fixes
  python tools/seo_title_optimizer.py --category wall_art  # audit one category
"""

import os, sys, json, re, argparse
from pathlib import Path
from collections import Counter

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

_env_path = ROOT / '.env'
if _env_path.exists():
    with open(_env_path) as f:
        for line in f:
            line = line.strip()
            if line and not line.startswith('#') and '=' in line:
                k, v = line.split('=', 1)
                os.environ.setdefault(k.strip(), v.strip())

# ── Style/Niche keyword banks for title optimization ──
STYLE_KEYWORDS = {
    'boho': ['boho', 'bohemian', 'boho chic', 'eclectic'],
    'farmhouse': ['farmhouse', 'rustic', 'country', 'vintage'],
    'nursery': ['nursery', 'kids room', 'baby room', 'playroom'],
    'kitchen': ['kitchen', 'dining room', 'cooking', 'recipe'],
    'minimalist': ['minimalist', 'modern', 'clean lines', 'simple'],
    'botanical': ['botanical', 'plant lover', 'greenery', 'garden'],
    'coastal': ['coastal', 'beach', 'ocean', 'nautical', 'seaside'],
    'kawaii': ['kawaii', 'cute', 'pastel', 'anime style'],
    'cottagecore': ['cottagecore', 'cottage', 'pastoral', 'whimsical'],
    'dark_academia': ['dark academia', 'moody', 'vintage library', 'scholarly'],
}

ROOM_KEYWORDS = [
    'living room', 'bedroom', 'bathroom', 'kitchen', 'nursery',
    'office', 'home office', 'dorm room', 'entryway', 'gallery wall',
]

INTENT_KEYWORDS = [
    'instant download', 'printable', 'digital download', 'wall art',
    'home decor', 'gift for', 'housewarming gift', 'birthday gift',
]

# ── Title Quality Scoring ──
def score_title(title: str) -> dict:
    """Score a listing title against 2026 Etsy SEO best practices."""
    issues = []
    score = 100

    # Rule 1: Length check (140 char max, 40 char first-segment)
    if len(title) < 40:
        issues.append("TOO SHORT: Title under 40 chars — missing keyword opportunities")
        score -= 20
    if len(title) > 140:
        issues.append("TOO LONG: Over 140 chars — Etsy truncates")
        score -= 10

    # Rule 2: First 40 characters should have a descriptive keyword
    first_40 = title[:40].lower()
    has_style = any(kw in first_40 for kwlist in STYLE_KEYWORDS.values() for kw in kwlist)
    has_room = any(kw in first_40 for kw in ROOM_KEYWORDS)
    has_intent = any(kw in first_40 for kw in ['printable', 'digital', '3d printed'])
    if not has_style and not has_room:
        issues.append("WEAK LEAD: First 40 chars lack a style/room keyword (mobile users only see this)")
        score -= 15

    # Rule 3: Generic title detection
    generic_patterns = [
        r'^(printable wall art|wall art print|instant download)',
        r', printable wall art, instant download$',
        r'^art print',
    ]
    for pat in generic_patterns:
        if re.search(pat, title.lower()):
            issues.append(f"GENERIC: Title follows a generic pattern — needs buyer-intent rewrite")
            score -= 20
            break

    # Rule 4: Uses pipe separators (better readability)
    if '|' not in title and ',' in title:
        issues.append("FORMAT: Consider pipe (|) separators instead of commas for readability")
        score -= 5

    # Rule 5: Missing "instant download" or "printable" for digital items
    lower = title.lower()
    if 'instant download' not in lower and 'digital download' not in lower and 'printable' not in lower and '3d printed' not in lower:
        issues.append("MISSING: No 'instant download' or 'printable' keyword — buyers filter by this")
        score -= 10

    # Rule 6: Keyword stuffing detection
    words = re.findall(r'\w+', title.lower())
    word_counts = Counter(words)
    repeated = {w: c for w, c in word_counts.items() if c >= 3 and w not in ('the', 'and', 'of', 'for', 'a', 'in')}
    if repeated:
        issues.append(f"STUFFED: Words repeated 3+ times: {repeated}")
        score -= 10

    return {
        'score': max(0, score),
        'grade': 'A' if score >= 85 else 'B' if score >= 70 else 'C' if score >= 55 else 'D' if score >= 40 else 'F',
        'issues': issues,
        'length': len(title),
    }


def score_tags(tags: list[str]) -> dict:
    """Score a listing's tags against best practices."""
    issues = []
    score = 100

    if len(tags) < 13:
        issues.append(f"MISSING TAGS: Only {len(tags)}/13 tags used — fill all 13 slots")
        score -= (13 - len(tags)) * 5

    single_word = [t for t in tags if ' ' not in t and len(t) > 0]
    if single_word:
        issues.append(f"SINGLE-WORD TAGS: {single_word[:3]} — use multi-word phrases instead")
        score -= len(single_word) * 3

    over_20 = [t for t in tags if len(t) > 20]
    if over_20:
        issues.append(f"OVER 20 CHARS: {over_20} — Etsy truncates tags over 20 characters")
        score -= len(over_20) * 5

    return {
        'score': max(0, score),
        'issues': issues,
        'tag_count': len(tags),
    }


# ── Optimized title suggestions by category ──
TITLE_TEMPLATES = {
    'wall_art': "{subject} Wall Art Print | {style} {room} Decor | Printable Instant Download | {mood}",
    'coloring_pages': "{subject} Coloring Pages | {count} Page Printable Coloring Book | Instant Download PDF | {audience}",
    'digital_planner': "{theme} Digital Planner {year} | {platform} iPad Planner | Fillable PDF | Instant Download",
    'paper_pack': "{theme} Digital Paper Pack | {count} Seamless Patterns | Scrapbook Paper | Instant Download",
    'sticker_pack': "{theme} Digital Sticker Pack | {platform} Planner Stickers | Instant Download PNG",
    'svg_bundle': "{theme} SVG Bundle | {count} Cut Files for Cricut Silhouette | Instant Download",
    '3d_print_physical': "3D Printed {item} | {style} {room} Decor | Custom Color | {feature}",
}


def generate_optimized_title(current_title: str, category: str) -> str:
    """Generate an SEO-optimized title suggestion based on the current title and category."""
    lower = current_title.lower()

    # Extract subject from current title
    # Remove common suffixes
    subject = current_title.split(',')[0].split('|')[0].strip()
    if subject.endswith('Print'):
        subject = subject[:-5].strip()
    if subject.endswith('Art'):
        subject = subject[:-3].strip()

    # Detect style
    style = 'Modern'
    for style_name, keywords in STYLE_KEYWORDS.items():
        if any(kw in lower for kw in keywords):
            style = style_name.replace('_', ' ').title()
            break

    # Detect room
    room = 'Home'
    for r in ROOM_KEYWORDS:
        if r in lower:
            room = r.title()
            break

    if category == 'wall_art':
        return f"{subject} Wall Art Print | {style} {room} Decor | Printable Digital Download"
    elif category == 'coloring_pages':
        return f"{subject} | Printable Coloring Book for Adults & Kids | Instant Download PDF"
    elif category == '3d_print_physical':
        return f"3D Printed {subject} | {style} Home Decor | Custom Color Choice | Unique Gift"
    elif category == 'paper_pack':
        return f"{subject} | Seamless Digital Paper Pack | Scrapbook Backgrounds | Instant Download"
    else:
        return f"{subject} | {style} Digital Download | Instant Access"


def audit_catalog(category_filter: str = None, show_fixes: bool = False):
    """Audit the full product catalog for SEO quality."""
    import openpyxl
    tracker_path = Path('/home/sjack7/Desktop/Business/OnBrandCraftz_Business_Tracker_2026-07-19.xlsx')
    if not tracker_path.exists():
        print("❌ Business tracker not found")
        return

    wb = openpyxl.load_workbook(tracker_path)
    ws = wb['Products']

    results = {'A': 0, 'B': 0, 'C': 0, 'D': 0, 'F': 0}
    all_issues = []

    divider = "─" * 72
    print(f"\n{divider}")
    print(f"  SEO TITLE AUDIT — OnBrandCraftz")
    print(f"  {divider}")

    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
        pid, name, cat, status, price, lid, updated, notes = row[:8]
        if status != 'active' or not name:
            continue
        if category_filter and cat != category_filter:
            continue

        result = score_title(name)
        results[result['grade']] += 1

        if result['grade'] in ('C', 'D', 'F') or show_fixes:
            print(f"\n  [{result['grade']}] {result['score']}/100 — {cat}")
            print(f"      {name[:90]}")
            for issue in result['issues']:
                print(f"      ⚠ {issue}")
                all_issues.append(issue.split(':')[0])

            if show_fixes:
                optimized = generate_optimized_title(name, cat or 'wall_art')
                print(f"      → {optimized[:120]}")

    print(f"\n{divider}")
    print(f"  GRADE DISTRIBUTION")
    total = sum(results.values())
    for grade in ['A', 'B', 'C', 'D', 'F']:
        bar = '█' * results[grade]
        pct = (results[grade] / total * 100) if total else 0
        print(f"    {grade}: {results[grade]:>3} ({pct:>4.1f}%) {bar}")

    print(f"\n  TOP ISSUES:")
    issue_counts = Counter(all_issues)
    for issue, count in issue_counts.most_common(5):
        print(f"    {count:>3}x {issue}")
    print(divider)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Audit & optimize Etsy listing titles")
    parser.add_argument("--fix", action="store_true", help="Show optimized title suggestions")
    parser.add_argument("--category", type=str, default=None, help="Filter by category")
    args = parser.parse_args()

    audit_catalog(category_filter=args.category, show_fixes=args.fix)
